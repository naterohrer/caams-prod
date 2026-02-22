import io
import re
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .. import models
from ..database import get_db
from ..auth import require_any
from ..engine.mapper import compute_results, compute_summary
from .assessments import get_recommendations

router = APIRouter(prefix="/assessments", tags=["export"])

STATUS_COLORS = {
    "covered": "00B050",     # green
    "partial": "FFC000",     # amber
    "not_covered": "FF0000", # red
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F4F4F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OVERRIDE_FILL = PatternFill(fill_type="solid", fgColor="E8F4FD")  # light blue for overridden rows


def _safe_filename(name: str) -> str:
    """Strip characters that are unsafe in Content-Disposition filenames."""
    return re.sub(r'[^\w\-.]', '_', name)


def _safe_cell(value: str) -> str:
    """
    Prevent CSV/Excel formula injection.
    Strings starting with =, +, -, @, tab, or CR are prefixed with a single
    quote, which tells Excel/LibreOffice to treat the cell as literal text.
    """
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


@router.get("/{assessment_id}/export")
def export_assessment(assessment_id: int, db: Session = Depends(get_db), _=Depends(require_any)):
    assessment = db.query(models.Assessment).filter(
        models.Assessment.id == assessment_id
    ).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    results = compute_results(assessment, db)
    summary = compute_summary(results)

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Assessment", _safe_cell(assessment.name)])
    ws_summary.append(["Framework", _safe_cell(f"{assessment.framework.name} {assessment.framework.version}")])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Controls", summary["total"]])
    ws_summary.append(["Covered", summary["covered"]])
    ws_summary.append(["Partial", summary["partial"]])
    ws_summary.append(["Not Covered", summary["not_covered"]])
    ws_summary.append(["Coverage %", f"{summary['coverage_pct']}%"])
    overridden_count = sum(1 for r in results if r.get("status_overridden"))
    if overridden_count:
        ws_summary.append(["Manual Overrides", overridden_count])
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40

    # ── Sheet 2: Coverage Report ──────────────────────────────────────────────
    ws = wb.create_sheet("Coverage Report")
    headers = [
        "Control ID", "Title", "Status", "Overridden?",
        "Owner", "Team", "Evidence Owner",
        "Covered By", "Missing Tags", "Evidence Needed",
        "Notes", "Evidence / Process Link",
        "Override Justification", "Override Expires",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for result in results:
        status_label = result["status"].replace("_", " ").title()
        overridden = result.get("status_overridden", False)
        expires = result.get("override_expires")
        ws.append([
            result["control_id"],
            result["title"],
            status_label,
            "Yes (compensating)" if overridden else "",
            _safe_cell(result.get("owner") or ""),
            _safe_cell(result.get("team") or ""),
            _safe_cell(result.get("evidence_owner") or ""),
            ", ".join(result["contributing_tools"]),
            ", ".join(result["missing_tags"]),
            "\n".join(result["evidence"]),
            _safe_cell(result.get("notes") or ""),
            _safe_cell(result.get("evidence_url") or ""),
            _safe_cell(result.get("override_justification") or ""),
            str(expires) if expires else "",
        ])
        row = ws.max_row
        status_cell = ws.cell(row=row, column=3)
        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=STATUS_COLORS.get(result["status"], "FFFFFF"),
        )
        status_cell.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=row, column=10).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical="top")
        # Highlight overridden rows subtly
        if overridden:
            for col in [1, 4, 13, 14]:
                ws.cell(row=row, column=col).fill = OVERRIDE_FILL

    col_widths = [12, 38, 14, 16, 22, 22, 22, 35, 28, 50, 40, 45, 40, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 20

    # ── Sheet 3: Evidence Checklist ───────────────────────────────────────────
    ws2 = wb.create_sheet("Evidence Checklist")
    checklist_headers = [
        "Control ID", "Control Title", "Evidence Item",
        "Status", "Owner", "Team", "Evidence Owner",
        "Notes", "Evidence / Process Link",
    ]
    ws2.append(checklist_headers)
    _style_header_row(ws2, len(checklist_headers))

    for result in results:
        status_label = result["status"].replace("_", " ").title()
        for evidence_item in result["evidence"]:
            ws2.append([
                result["control_id"],
                result["title"],
                evidence_item,
                status_label,
                _safe_cell(result.get("owner") or ""),
                _safe_cell(result.get("team") or ""),
                _safe_cell(result.get("evidence_owner") or ""),
                _safe_cell(result.get("notes") or ""),
                _safe_cell(result.get("evidence_url") or ""),
            ])

    checklist_col_widths = [12, 38, 45, 14, 22, 22, 22, 40, 45]
    for i, width in enumerate(checklist_col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 4: Recommendations ──────────────────────────────────────────────
    recs = get_recommendations(assessment_id, db)

    if recs:
        ws3 = wb.create_sheet("Recommendations")
        rec_headers = ["Capability Gap", "Controls Requiring It", "Control IDs"]
        ws3.append(rec_headers)
        _style_header_row(ws3, len(rec_headers))
        for rec in recs:
            ws3.append([
                rec.capability,
                rec.controls_count,
                ", ".join(rec.controls_detail),
            ])
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"assessment_{assessment_id}_{_safe_filename(assessment.name)}_coverage.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
