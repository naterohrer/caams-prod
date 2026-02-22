"""
CIS Controls xlsx → CAAMS framework JSON importer.

Handles the official CIS Controls spreadsheet format.  Column positions are
detected from the header row so the parser is resilient to minor layout
changes between versions (v8, v8.1, future releases, etc.).

Expected sheet columns (case-insensitive, partial-match):
  "control"      → top-level control number (1, 2, … 18)
  "safeguard"    → sub-control number (1.1, 1.2, …)
  "asset type"   → Devices | Applications | Network | Data | Users | N/A
  "security fun" → Identify | Protect | Detect | Respond | Recover
  "title"        → safeguard title
  "description"  → safeguard description
  "ig1" / "ig2" / "ig3" → Implementation Group membership

Usage:
    from app.importers.cis_xlsx import parse_cis_xlsx
    framework_json = parse_cis_xlsx("CIS_Controls_v8.xlsx", version="v8")
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


# ---------------------------------------------------------------------------
# Tag heuristics — derived from Asset Type + Security Function columns.
# These are reasonable starting points; operators should refine them.
# ---------------------------------------------------------------------------

_ASSET_REQUIRED: dict[str, list[str]] = {
    "devices":       ["asset-inventory", "asset-discovery"],
    "applications":  ["software-inventory"],
    "network":       ["network-monitoring"],
    "data":          ["data-protection"],
    "users":         ["identity", "access-management"],
}

_ASSET_OPTIONAL: dict[str, list[str]] = {
    "devices":       ["cmdb", "mdm", "network-monitoring"],
    "applications":  ["application-allowlisting", "patch-management", "cmdb"],
    "network":       ["firewall", "ids-ips", "network-segmentation"],
    "data":          ["encryption", "dlp", "backup"],
    "users":         ["mfa", "pam", "sso"],
}

_FUNCTION_REQUIRED: dict[str, list[str]] = {
    "identify":  ["asset-inventory"],
    "protect":   ["access-control"],
    "detect":    ["log-management", "monitoring"],
    "respond":   ["incident-response"],
    "recover":   ["backup"],
}

_FUNCTION_OPTIONAL: dict[str, list[str]] = {
    "identify":  ["vulnerability-scanning"],
    "protect":   ["hardening", "encryption"],
    "detect":    ["siem", "edr"],
    "respond":   ["forensics"],
    "recover":   ["backup-testing", "disaster-recovery"],
}


def _tags_for(asset_type: str, security_function: str) -> tuple[list[str], list[str]]:
    """Return (required_tags, optional_tags) for a safeguard row."""
    at = asset_type.lower().strip()
    sf = security_function.lower().strip()

    # Match prefixes so "Devices" matches "devices" and also "device"
    def _match(mapping: dict, key: str) -> list[str]:
        for k, v in mapping.items():
            if key.startswith(k) or k.startswith(key):
                return v
        return []

    req = list(dict.fromkeys(_match(_ASSET_REQUIRED, at) + _match(_FUNCTION_REQUIRED, sf)))
    opt = list(dict.fromkeys(_match(_ASSET_OPTIONAL, at) + _match(_FUNCTION_OPTIONAL, sf)))
    # Remove anything already in req from opt
    opt = [t for t in opt if t not in req]
    return req, opt


# ---------------------------------------------------------------------------
# CIS v8 top-level control titles as a fallback when the xlsx only contains
# safeguard-level rows without an explicit parent row.
# ---------------------------------------------------------------------------

_CIS_CONTROL_TITLES: dict[int, str] = {
    1:  "Inventory and Control of Enterprise Assets",
    2:  "Inventory and Control of Software Assets",
    3:  "Data Protection",
    4:  "Secure Configuration of Enterprise Assets and Software",
    5:  "Account Management",
    6:  "Access Control Management",
    7:  "Continuous Vulnerability Management",
    8:  "Audit Log Management",
    9:  "Email and Web Browser Protections",
    10: "Malware Defenses",
    11: "Data Recovery",
    12: "Network Infrastructure Management",
    13: "Network Monitoring and Defense",
    14: "Security Awareness and Skills Training",
    15: "Service Provider Management",
    16: "Application Software Security",
    17: "Incident Response Management",
    18: "Penetration Testing",
}


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    Scan rows until we find one that looks like a header (contains 'safeguard'
    or 'control' and 'title').  Return (row_index, col_map) where col_map maps
    logical field names to 0-based column indices.
    """
    KEYWORDS = {
        "control":           "ctrl_num",
        "safeguard":         "safeguard",
        "asset type":        "asset_type",
        "asset":             "asset_type",
        "security function": "sec_func",
        "security fun":      "sec_func",
        "function":          "sec_func",
        "title":             "title",
        "description":       "description",
        "ig1":               "ig1",
        "ig2":               "ig2",
        "ig3":               "ig3",
    }

    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).lower().strip()
            for kw, field in KEYWORDS.items():
                if kw in cell_str and field not in col_map:
                    col_map[field] = col_idx
                    break

        # Require at least safeguard + title to proceed
        if "safeguard" in col_map and "title" in col_map:
            return row_idx, col_map

    raise ValueError(
        "Could not find a header row containing 'safeguard' and 'title'. "
        "Check that you're pointing at the correct sheet."
    )


def _pick_sheet(wb, sheet_name: str | None):
    """Return the target worksheet, preferring sheet_name if given."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        return wb[sheet_name]

    # Try common CIS sheet names first
    for candidate in ("CIS Controls", "Controls", "CIS Safeguards", "Safeguards"):
        if candidate in wb.sheetnames:
            return wb[candidate]

    return wb.active  # fall back to first sheet


def _cell_val(row: tuple, idx: int | None) -> str:
    """Safely retrieve a cell value as a stripped string."""
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _is_integer(s: str) -> bool:
    try:
        int(s)
        return True
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# Main parse function
# ---------------------------------------------------------------------------

def parse_cis_xlsx(
    xlsx_path: str | Path,
    *,
    version: str = "v8",
    sheet_name: str | None = None,
    min_ig: int = 1,
) -> dict[str, Any]:
    """
    Parse a CIS Controls xlsx file and return a CAAMS-compatible framework dict.

    Parameters
    ----------
    xlsx_path   : path to the .xlsx file
    version     : version string written into the JSON (e.g. "v8", "v8.1")
    sheet_name  : explicit worksheet name; auto-detected if omitted
    min_ig      : only include safeguards tagged for this IG or higher
                  (1 = all safeguards, 2 = IG2+, 3 = IG3 only)
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = _pick_sheet(wb, sheet_name)

    header_row_idx, col = _find_header_row(ws)

    # Unpack column indices (None if column was not found in the header)
    c_ctrl     = col.get("ctrl_num")
    c_safe     = col.get("safeguard")
    c_asset    = col.get("asset_type")
    c_func     = col.get("sec_func")
    c_title    = col.get("title")
    c_desc     = col.get("description")
    c_ig2      = col.get("ig2")
    c_ig3      = col.get("ig3")

    # ── Group safeguard rows by top-level control number ──────────────────────
    # Each key is an int (1–18).  Value is a list of safeguard dicts.
    controls_raw: dict[int, dict] = {}   # ctrl_num → {title, description, safeguards}

    all_rows = list(ws.iter_rows(values_only=True))

    for row in all_rows[header_row_idx + 1:]:
        safeguard = _cell_val(row, c_safe)
        if not safeguard:
            continue

        # Parse control number from safeguard like "1.1" → 1
        # Some xlsx files put the integer control number in the first column.
        ctrl_raw = _cell_val(row, c_ctrl) if c_ctrl is not None else ""

        try:
            if "." in safeguard:
                ctrl_num = int(safeguard.split(".")[0])
            elif _is_integer(safeguard):
                # This is a control-level row (the parent), not a safeguard
                ctrl_num = int(safeguard)
                ctrl_title = _cell_val(row, c_title)
                ctrl_desc  = _cell_val(row, c_desc)
                if ctrl_num not in controls_raw:
                    controls_raw[ctrl_num] = {
                        "title": ctrl_title,
                        "description": ctrl_desc,
                        "safeguards": [],
                    }
                elif not controls_raw[ctrl_num].get("title"):
                    controls_raw[ctrl_num]["title"] = ctrl_title
                    controls_raw[ctrl_num]["description"] = ctrl_desc
                continue
            elif ctrl_raw and _is_integer(ctrl_raw):
                ctrl_num = int(ctrl_raw)
                # If no "." in safeguard and ctrl_raw exists, this might be
                # a control-level row in some formats
                if not re.match(r"^\d+\.\d", safeguard):
                    ctrl_title = _cell_val(row, c_title)
                    ctrl_desc  = _cell_val(row, c_desc)
                    if ctrl_num not in controls_raw:
                        controls_raw[ctrl_num] = {
                            "title": ctrl_title,
                            "description": ctrl_desc,
                            "safeguards": [],
                        }
                    continue
            else:
                continue
        except (ValueError, IndexError):
            continue

        # IG filtering
        if min_ig >= 2:
            ig2_val = _cell_val(row, c_ig2)
            ig3_val = _cell_val(row, c_ig3)
            in_ig2 = bool(ig2_val and ig2_val.lower() not in ("", "n/a", "no", "false"))
            in_ig3 = bool(ig3_val and ig3_val.lower() not in ("", "n/a", "no", "false"))
            if min_ig == 2 and not (in_ig2 or in_ig3):
                continue
            if min_ig == 3 and not in_ig3:
                continue

        title = _cell_val(row, c_title)
        desc  = _cell_val(row, c_desc)
        asset = _cell_val(row, c_asset)
        func  = _cell_val(row, c_func)

        if ctrl_num not in controls_raw:
            controls_raw[ctrl_num] = {"title": "", "description": "", "safeguards": []}

        controls_raw[ctrl_num]["safeguards"].append({
            "id":          safeguard,
            "title":       title,
            "description": desc,
            "asset_type":  asset,
            "sec_func":    func,
        })

    # ── Build CAAMS control objects ───────────────────────────────────────────
    controls_out: list[dict] = []

    for ctrl_num in sorted(controls_raw.keys()):
        raw = controls_raw[ctrl_num]
        safeguards = raw.get("safeguards", [])

        # Derive control-level title/description
        ctrl_title = (
            raw.get("title")
            or _CIS_CONTROL_TITLES.get(ctrl_num, f"Control {ctrl_num}")
        )
        ctrl_desc = (
            raw.get("description")
            or (safeguards[0]["description"] if safeguards else "")
        )

        # Aggregate tags across all safeguards in this control
        all_req: list[str] = []
        all_opt: list[str] = []
        for sg in safeguards:
            req, opt = _tags_for(sg["asset_type"], sg["sec_func"])
            all_req.extend(req)
            all_opt.extend(opt)

        # Deduplicate while preserving order
        required_tags = list(dict.fromkeys(all_req))
        optional_tags = list(dict.fromkeys(
            t for t in all_opt if t not in required_tags
        ))

        sub_controls = [
            {"id": sg["id"], "title": sg["title"]}
            for sg in safeguards
            if sg.get("title")
        ]

        controls_out.append({
            "control_id":    f"CIS-{ctrl_num}",
            "title":         ctrl_title,
            "description":   ctrl_desc,
            "required_tags": required_tags,
            "optional_tags": optional_tags,
            "evidence":      [],        # operators fill these in
            "sub_controls":  sub_controls,
        })

    if not controls_out:
        raise ValueError(
            "No controls were parsed.  Check that the sheet and column format "
            "match the expected CIS Controls xlsx layout."
        )

    return {
        "name":     "CIS Controls",
        "version":  version,
        "controls": controls_out,
    }
